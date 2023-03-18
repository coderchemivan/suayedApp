import csv
import datetime
import chardet
import pandas as pd
import xlwings as xw
from openpyxl import load_workbook
import calendar
import mysql.connector
import re

class DB_admin():

    def __init__(self, archivo_materias=None,usuario=None):
        self.archivo_materias = archivo_materias
        self.usuario = usuario
        self.db_connection()

    def db_connection(self):
        self.conn = mysql.connector.connect(user="root", password="123456",
                                       host="localhost",
                                       database="fca_materias",
                                       port='3306'
                                       )
        self.cur = self.conn.cursor()

    def semester_añadido_info(self,semestre=None,accion=None,materias=None):
        if accion == 'insertar':
            try:
                self.cur.execute("INSERT INTO semestres (name) VALUES ('{}')".format(semestre))
                self.conn.commit()
            except  Exception as e:
                print(e)
        elif accion == 'update':
            self.cur.execute("UPDATE user_settings SET semestre_a_seleccionar = {} WHERE user_id = 1".format(semestre))
            self.conn.commit()
        elif accion == 'select':
            self.cur.execute("SELECT semestre_a_seleccionar FROM user_settings")
            semestre = self.cur.fetchone()[0]
            return semestre
        elif accion == 'insertar_materias':
            semestre_id = self.cur.execute("SELECT id FROM semestres WHERE name = '{}'".format(semestre))
            semestre_id = self.cur.fetchone()[0]
            for materia in materias:
                clave = materia['clave_materia']
                grupo = materia['grupo']
                self.cur.execute("INSERT INTO materias_usuario (clave_materia,user_id,semestre_id,grupo,meta) VALUES ({},{},{},{},{})".format(clave,self.usuario,semestre_id,grupo,10))
                self.conn.commit()

            
    def materias_por_semestre(self,semestre,usuario=None,carrera=None):
        if usuario == None:
            self.cur.execute("SELECT*FROM materias_fca WHERE semestre = {} and {} = 1;".format(semestre,carrera))
            materias = self.cur.fetchall()
            materias = [dict(zip(['clave','materia','semestre','Administración'],materia)) for materia in materias]
        else:
            semestre = 'Semestre ' + semestre
            self.cur.execute(f'''
                            select distinct(materia),clave_materia,grupo from 
                            (select mf.materia,mu.clave_materia,mu.grupo from materias_usuario as mu
                            left join materias_fca as mf 
                            ON mu.clave_materia = mf.clave_materia
                            where mu.semestre_id= (select id from semestres where name = '{semestre}')) as tabla;
            ''')
            materias = self.cur.fetchall()
            materias = [dict(zip(['materia','clave_materia','grupo'],materia)) for materia in materias]
        return materias

    def load_data(self,archivo = None,semestre=None):
        semestre_id = self.cur.execute("SELECT id FROM semestres WHERE name = '{}'".format(f"Semestre {semestre}"))
        semestre_id = self.cur.fetchone()[0]
        df_actividades = pd.read_csv(archivo,encoding='latin1')
        for row in df_actividades.itertuples():
            fecha_entrega = row.fecha_entrega.split('/')
            fecha_entrega = fecha_entrega[2] + '-' + fecha_entrega[1] + '-' + fecha_entrega[0]
            ponderacion = row.valor
            self.cur.execute('''REPLACE INTO actividades (name,semestre,clave_materia,usuario,fecha_entrega,valor) 
                                VALUES ('{}',{},{},{},'{}',{})'''.format(row.name,semestre_id,row.clave_materia,self.usuario,fecha_entrega,ponderacion))
            self.conn.commit()

    def lista_semester(self):  # Devuelve una lista con los semestres disponibles

        self.cur.execute(
            '''
                SELECT DISTINCT name FROM semestres
            '''
        )
        semestres = self.cur.fetchall()
        semestres_ = list()
        for semestre in semestres:
            semestres_.append(semestre[0])
        return semestres_

    def fetch_semester_selected_info(self):
        semestre_info = {}
        name= self.cur.execute('''
            select name from semestres where id = 
            (SELECT semestre_sel FROM user_settings);
        ''')
        name = self.cur.fetchone()[0]
        id = self.cur.execute('''SELECT semestre_sel FROM user_settings''')
        id = self.cur.fetchone()[0]
        semestre_info['name'] = name
        semestre_info['id'] = id
        return semestre_info

    def define_semester(self, semestre, archivo_aux=None):
        '''SET THE SEMESTER WE'RE GONNA WORK WITH IN USER_SETTINGS TABLE'''
        self.cur.execute("SELECT id FROM semestres WHERE name = '" + semestre + "'")  ## busca el id del semestre que se seleccionó
        semestre = self.cur.fetchone()[0]

        self.cur.execute('''
            UPDATE user_settings SET tipo_lista = 'TwoLineRightIconListItem' WHERE user_id = 1 
        ''')
        self.cur.execute('''
            UPDATE user_settings SET status = 'Por entregar' WHERE user_id = 1
        ''')
        self.cur.execute("UPDATE user_settings SET semestre_sel = '" + str(semestre) + "' WHERE user_id = 1") ## establece el semestre seleccionado en la tabla user_settings
        subject_name = self.obtener_materia_name_(modo=2,semestre_=semestre)
        self.cur.execute("UPDATE user_settings SET materia_sel = '" + subject_name[0] + "' WHERE user_id = 1") ## establece la primera materia econtrada del semestre establecido
        self.conn.commit()

    def dias_con_pendientes(self, modo, month=None, año=None, fecha=None):
        if modo == 1:  ## Regresa una lista con los días que tienen pendientes
            primer_dia_mes = año + '-' + str(month).zfill(2) +'-' + '01'
            ultimo_dia_mes = año + '-' + str(month).zfill(2) + '-' + str(calendar.monthrange(int(año), month)[1])

            fechas_mes = "SELECT fecha_entrega FROM actividades " \
                                  "WHERE fecha_entrega BETWEEN '" + primer_dia_mes + "' AND '" + ultimo_dia_mes + "'"

            self.cur.execute(fechas_mes)
            fechas_ = self.cur.fetchall()
            lista_dias = list()
            if len(fechas_)!=0:
                fechas,= list(zip(*fechas_))
                for fecha in fechas:
                    fecha = fecha.split("-")[2]
                    lista_dias.append(fecha)

            return lista_dias

        elif modo ==2:  ## Regresa una lista con el nomnbre de las materias, act y status
            dia = "SELECT name,clave_materia FROM actividades " \
                                  "WHERE fecha_entrega = '" + fecha + "'"
            self.cur.execute(dia)
            registros = self.cur.fetchall()
            materias = list()
            actividades = list()
            status = list()
            materia_actividad = list()
            for registro in registros:
                clave = registro[1]
                acti_name = registro[0]
                acti_name = self.encode_decode_activity(acti_name)
                actividades.append(acti_name)
                materia_nombre = '''SELECT materia FROM materias_fca  
                                    JOIN actividades 
                                    ON materias_fca.clave_materia = actividades.clave_materia
                                    where materias_fca.clave_materia =''' + str(clave) + ''' LIMIT 1'''
                print(materia_nombre)
                self.cur.execute(materia_nombre)
                status_ = self.cur.fetchall()
                status_,= list(zip(*status_))
                for stat in status_:
                    status.append(stat)
                
                self.cur.execute("SELECT status FROM actividades where clave_materia ='" + str(clave) + "' AND name= '" + str(registro[0]) + "'")
                materia_nombre = self.cur.fetchall()
                for materia in materia_nombre:
                    materias.append(materia[0])
                
            materia_actividad.append(materias)
            materia_actividad.append(actividades)
            materia_actividad.append(status)
            return materia_actividad

    def remember_status(self,modo):  ##Esta función juega con el remember_status
        remember_status = ('''
            SELECT recordar_inicio_sesion FROM user_settings WHERE user_id = 1
        '''
        )
        self.cur.execute(remember_status)
        remember_status = self.cur.fetchone()[0]

        if modo ==1: ## Se presionó el butón


            if remember_status == 0:
                self.cur.execute('''
                    UPDATE user_settings SET recordar_inicio_sesion=1 WHERE user_id = 1
                ''')
            else:
                self.cur.execute('''
                    UPDATE user_settings SET recordar_inicio_sesion=0 WHERE user_id = 1
                ''')

            self.conn.commit()
        else: #pura consulta
            return remember_status


    def extracción_materias(self):
        '''Devuelve las materias conforme al semestre seleccionado'''
        semestre = "SELECT semestre_sel FROM user_settings WHERE user_id = 1"
        self.cur.execute(semestre)
        semestre = self.cur.fetchone()[0]


        materias = "SELECT DISTINCT materia FROM materias_fca JOIN actividades ON materias_fca.clave_materia = actividades.clave_materia WHERE actividades.semestre ='" + str(semestre) + "'"
        self.cur.execute(materias)
        materias = self.cur.fetchall()
        materias_ = list()
        for materia in materias:
            materias_.append(materia[0])
        return materias_

    def status(self,modo,materia=None,status_=None):
        semestre_id = self.fetch_semester_selected_info()['id']
        if modo == 1:  ## Actualiza solo el estado
            act_status = self.cur.execute("SELECT name,clave_materia,date_format(fecha_entrega,'%d-%m-%Y'),date_format(entregada_el,'%d-%m-%Y'),status FROM actividades where semestre = {}".format(semestre_id))
            act_status = [list(i) for i in self.cur.fetchall()]

            i = 0
            for status in act_status:
                act = status[0]
                clave_materia = status[1]
                fecha_entrega = status[2]
                entregada_el = status[3]
                fecha_entrega = datetime.datetime.strptime(fecha_entrega, '%d-%m-%Y')
                hoy = datetime.datetime.today()
                if entregada_el == None:
                    status_ = 'Por entregar' if hoy.date() <= fecha_entrega.date() else 'Atrasada'
                else:
                    entregada_el = datetime.datetime.strptime(entregada_el, '%d-%m-%Y')
                    status_ = 'Entregada a tiempo' if entregada_el.date()<=fecha_entrega.date() else 'Entregada con atraso'
                i+=1
                self.cur.execute("UPDATE actividades SET status = '{}' WHERE name = '{}' and clave_materia = {}".format(status_,act,clave_materia))
                self.conn.commit()
        elif modo ==2: ## Trae las actividaddes de las materias con el estado dado
            act_estatus = self.cur.execute(
                "SELECT name,date_format(fecha_entrega,'%d-%m-%Y'),date_format(entregada_el,'%d-%m-%Y'),status FROM actividades WHERE clave_materia= '" + str(materia) + "' AND status = '" + status_ +"'")
            act_status = [list(i) for i in self.cur.fetchall()]
            activity_date = dict()
            for act in act_status:
                actividad = self.encode_decode_activity(act[0])
                fecha_entrega = act[1]
                entregada_el = act[2]
                fecha_entrega = datetime.datetime.strptime(fecha_entrega, '%d-%m-%Y')
                if status_ == 'Por entregar':
                    activity_date[actividad] = f'Due to {fecha_entrega.date().strftime("%d-%m-%Y")}'
                    self.cur.execute('''
                        UPDATE user_settings SET tipo_lista = 'TwoLineRightIconListItem' WHERE user_id = 1
                    ''')
                elif status_ == 'Entregada a tiempo':
                    self.cur.execute('''
                        UPDATE user_settings SET tipo_lista = 'TwoLineListItem' WHERE user_id = 1
                    ''')
                    entregada_el = datetime.datetime.strptime(entregada_el, '%d-%m-%Y')
                    activity_date[actividad] = f'Delivered {(datetime.datetime.today()-entregada_el).days} days ago'
                elif status_ == 'Entregada con atraso':
                    self.cur.execute('''
                        UPDATE user_settings SET tipo_lista = 'TwoLineListItem' WHERE user_id = 1
                    ''')
                    entregada_el = datetime.datetime.strptime(entregada_el, '%d-%m-%Y')
                    activity_date[actividad] = f'Delivered {(datetime.datetime.today()-entregada_el).days} days ago'
                elif status_ == 'Atrasada':
                    self.cur.execute('''
                        UPDATE user_settings SET tipo_lista = 'TwoLineRightIconListItem' WHERE user_id = 1
                    ''')
                    #tranform fecha_entrega to %d-%b-%Y format
                    fecha_entrega = fecha_entrega.date().strftime('%d-%b-%y')
                    activity_date[actividad] = f'This activity was due for {(fecha_entrega)}'
            self.conn.commit()
            return activity_date
        elif modo ==3: ## Busca cual es el estado en user_settings
            self.cur.execute("SELECT status FROM user_settings WHERE user_id = 1")
            status = self.cur.fetchone()[0]
            return status

        elif modo == 4: ## Cambia el estado en user_setting según lo seleccionado
            self.cur.execute('''
                UPDATE user_settings SET status = "'''     + status_  + '''" WHERE user_id = 1
            ''')
            self.conn.commit()

        elif modo ==5:
            actividades = self.cur.execute(
                "SELECT name FROM actividades WHERE clave_materia= '" + str(materia) + "'")

            actividades = [i[0] for i in self.cur.fetchall()]
            return actividades
    def encode_decode_activity(self,act,abbrv=True,decode=True):
        if decode == True:
            act = str(act)
            tipos_actividades = {'1': 'Actividad', '2': 'Act.complementaria','3': 'Cuestionario_refor','4': 'Act. lo que aprendí', '5': 'Foro','6': 'Act. integradora','7': 'Examen'} if abbrv == True else {'1': 'Actividad', '2': 'Actividad complementaria','3': 'Cuestionario de reforzamiento','4': 'Lo que aprendí', '5': 'Foro','6': 'Actividad integradora','7': 'Examen'}
            unidad = act[0:2] if len(act) == 4 and act[0:2] != '99' else act[0]
            activity_code = act[2] if len(act) == 4 else act[1]
            activity_num = act[-1]
            activity_num = "" if activity_num == "0" else activity_num
            if  abbrv == True:
                act = f'U{unidad} {tipos_actividades[activity_code]} {activity_num}' if act[0:2] != '99' else f'{tipos_actividades[activity_code]}'
            else:
                act = f'Unidad {unidad} / {tipos_actividades[activity_code]} {activity_num} / ' if activity_num != "" else f'Unidad {unidad} / {tipos_actividades[activity_code]} / '
        else:
            act = act.replace('U','').replace('Actividad','1').replace('Act.complementaria','2').replace('Cuestionario_refor','3').replace('Act. lo que aprendí','4').replace('Foro','5').replace('Act. integradora','6').replace('Examen','7') 
            act = act.replace(' ','')
            act = act + '0' if len(act) == 2 else act
        return act

    def define_activity(self, actividad):
        '''Estblece en la tabla user_settings el valor de la act seleccionada en la pantalla 2'''
        actividad = self.encode_decode_activity(actividad,abbrv=False,decode=False)
        self.cur.execute('''
            UPDATE user_settings SET act_sel = "''' + actividad + '''" WHERE user_id = 1
        ''')
        self.conn.commit()

    def estado_actividad(self, clave, actividad):
        actividad = self.encode_decode_activity(actividad,abbrv=False,decode=False)
        '''Busca los valores del feedback de la actividad seleccionada en la p2 y los muestra en la pantalla de feedback'''
        self.cur.execute(
            "SELECT date_format(entregada_el,'%d-%m-%Y'),valor,status,calificacion,calificada_en,comentarios FROM actividades WHERE clave_materia= '" + str(clave) + "' AND name = '" + actividad + "'")
        feedback = self.cur.fetchone()
        feedback_ = list()
        feedback = [elem if elem != None else "" for elem in feedback]
        entregada_el = feedback[0]
        valor = feedback[1]
        status = feedback[2]
        cal = feedback[3]
        calificada_en = feedback[4]
        comentarios = feedback[5]
        feedback_.append(entregada_el)
        feedback_.append(valor)
        feedback_.append(status)
        feedback_.append(cal)
        feedback_.append(calificada_en)
        feedback_.append(comentarios)
        return feedback_

    def obtener_materia_name_(self,modo,subject_name_=None,semestre_=None,clave_materia= None):
        if modo ==1: ## Busca el nombre de la materia
            self.cur.execute("SELECT materia_sel FROM user_settings WHERE user_id = 1")
            subject_name = self.cur.fetchone()
            return subject_name
        elif modo ==2: ## Seleccionad el primer nombre de la materia que encuentre de acuerdo al semestre
            self.cur.execute( '''SELECT DISTINCT materia FROM materias_fca
                                JOIN actividades  
                                ON materias_fca.clave_materia = actividades.clave_materia   
                                WHERE actividades.semestre = {}
                                LIMIT 1'''.format(semestre_))
            subject_name = self.cur.fetchone()
            return subject_name
        elif modo ==3:   ## Busca la clave de la materia
            self.cur.execute( '''
            SELECT DISTINCT mf.clave_materia FROM materias_fca as mf
            JOIN actividades as act
            ON mf.clave_materia = act.clave_materia
            WHERE materia = "''' + subject_name_ + '''"''')
            clave = self.cur.fetchone()
            return clave

        elif modo == 4: ## Busca el nombre de la actividdad seleccionada en la pantalla 2
            self.cur.execute("SELECT act_sel FROM user_settings WHERE user_id = 1")
            acti_name = self.cur.fetchone()
            acti_name = acti_name[0]
            acti_name = acti_name.replace(" / ", "_").replace("Unidad ", "U").replace("/", "").replace(
                "Actividad complementaria", "Act_compl"). \
                replace("Cuestionario de reforzamiento", "Cuest_refor").replace("Actividad", "Act")
            return acti_name
        elif modo == 5:
            self.cur.execute("SELECT materia FROM materias_fca WHERE clave_materia={} ".format(clave_materia))
            subject_name = self.cur.fetchone()[0]
            return subject_name
        elif modo == 6:
            self.cur.execute("SELECT clave_materia FROM materias_fca WHERE materia='{}' ".format(subject_name_))
            clave_materia = self.cur.fetchone()[0]
            return clave_materia


    def list_type(self):
        self.cur.execute("SELECT tipo_lista FROM user_settings WHERE user_id = 1")
        tipo_lista = self.cur.fetchone()
        return tipo_lista[0]

    def obtener_materia_grupo(self, clave):
        self.cur.execute("SELECT  grupo FROM materias_usuario WHERE clave_materia= '" + str(
            clave)  + "'")
        grupo = self.cur.fetchone()[0]
        return grupo

    def actualizar_DB(self, materia, actividad):
        actividad = self.encode_decode_activity(actividad,decode=False)
        self.cur.execute("SELECT date_format(fecha_entrega,'%d-%m-%Y') FROM actividades WHERE clave_materia= '" + str(
            materia) + "' AND name = '" + actividad + "'")
        datos_act = self.cur.fetchall()

        hoy = datetime.datetime.today()
        self.cur.execute("UPDATE actividades SET entregada_el = '" + datetime.datetime.strftime(hoy,"%Y-%m-%d") + "' WHERE clave_materia= '" + str(
            materia) + "' AND name = '" + actividad + "'")

        fecha_entrega = datos_act[0][0]
        fecha_entrega = datetime.datetime.strptime(fecha_entrega,"%d-%m-%Y")
        if hoy > fecha_entrega:
            estado = "Entregada con atraso"
        else:
            estado = "Entregada a tiempo"
        self.cur.execute("UPDATE actividades SET status = '" + estado + "' WHERE clave_materia= '" + str(
            materia) + "' AND name = '" + actividad + "'")
        self.conn.commit()
    
    def update_fecha_entregada(self, fecha_entregada):
        materia_act = self.cur.execute("SELECT materia_sel,act_sel FROM user_settings WHERE user_id = 1")
        materia_act = self.cur.fetchall()
        
        materia = materia_act[0][0]
        actividad = materia_act[0][1]
        clave_materia = self.cur.execute("SELECT clave_materia FROM materias_fca WHERE materia = '{}'".format(materia))
        clave_materia = self.cur.fetchall()
        clave_materia = clave_materia[0][0]
        self.cur.execute("UPDATE actividades SET entregada_el = date_format('{}','%Y-%m-%d') WHERE clave_materia= '{}' AND name = '{}'".format(fecha_entregada,clave_materia,actividad))
        self.conn.commit()

    def condensado_tareas(self,clave):
        semestre =self.semestre_seleccionado()
        act_valor_cali = list()
        d = ("SELECT name,valor,calificacion FROM actividades WHERE usuario = 1 AND semestre = '{}' AND clave_materia = '{}'").format(str(semestre),str(clave))
        self.cur.execute(d)
        table_rows = self.cur.fetchall()
        df = pd.read_sql(d,con = self.conn)

        char_to_replace = {" / ":"_",
                           "Unidad ": "U",
                           "Actividad complementaria": "Act_compl",
                           "Cuestionario de reforzamiento": "Cuest_refor",
                            "Actividad": "Act",
                            "Actividad integradora": "Act_integradora",
                            "/": ""}
        for key,value in char_to_replace.items():
            df['name'] = df['name'].str.replace(key,value)
        df['valor'] = df['valor'].apply(lambda x: float(x)*100)
        df['valor'] = df['valor'].apply(lambda x: round(x,1))
        df['valor'] = df['valor'].apply(lambda x: str(x) + "%")
        

        #reemplazar nan por "pendiente"
        df['calificacion'] = df['calificacion'].fillna("Pendiente")
        df.rename(columns={'name':'Act','valor':'Valor','calificacion':'Calificación'},inplace=True)
        return df

    def semestre_seleccionado(self):
        semestre = "SELECT semestre_sel FROM user_settings WHERE user_id = 1"
        self.cur.execute(semestre)
        semestre = self.cur.fetchone()[0]
        return semestre

    def user_info(self,columnas):
        columnas = ",".join(columnas)
        user_info = self.cur.execute("SELECT {} FROM users WHERE id = 1".format(columnas))
        datos_usuario = self.cur.fetchall()
        #convertir datos_usuario a diccionario
        datos_usuario = dict(zip(columnas.split(","),datos_usuario[0]))
       
       
        
        return datos_usuario
        

class estatus_feedback():
    def __init__(self, archivo_materias, materia, actividad):
        self.archivo_materias = archivo_materias
        self.actividad = actividad
        self.materia = materia

    def resultados(self):
        with open(self.archivo_materias, 'r') as file:
            reader = csv.reader(file)
            activity_status = list()
            for line in reader:

                if line[1] == self.materia and line[2] == self.actividad:
                    fecha_entregada = line[4]
                    status = line[5]
                    calificacion = line[6]
                    calificada_el = line[7]
                    comentarios = line[8]
                    valor = line[13]

                    activity_status.append(fecha_entregada)
                    activity_status.append(status)
                    activity_status.append(calificacion)
                    activity_status.append(calificada_el)
                    activity_status.append(comentarios)
                    activity_status.append(valor)
                    break

            return activity_status


class vaciar_feedback():
    def __init__(self, archivo_materias, materia, actividades_feedback):
        self.archivo_materias = archivo_materias
        self.materia = materia
        self.actividades_feedback = actividades_feedback

    def vaciar_resultados(self):
        with open(self.archivo_materias, 'r') as file:

            reader = csv.reader(file)
            myList = list(reader)
            for line in reader:
                myList.append(line)
            for k, v in self.actividades_feedback.items():
                c = 0
                for row in myList:
                    if row[3] == self.materia and row[4] == k:
                        if v[0]!="" and v[0]!="Sin envío":
                            v[0] = float(v[0].split("/")[0])
                        else:pass
                        myList[c][8] = v[0]  # calificacion
                        myList[c][9] = v[1]  # calificada el
                        myList[c][10] = v[2]  # comentarios
                    c += 1
                    continue

                with open(self.archivo_materias, 'rb') as rawdata:
                    result = chardet.detect(rawdata.read(100000))
                my_new_list = open(self.archivo_materias, 'w', newline='', encoding=result['encoding'])
                csv_writer = csv.writer(my_new_list)
                for line in myList:
                    try:
                        csv_writer.writerow(line)
                    except:
                        line[10] = line[10].encode('utf8').decode('ascii', 'ignore')
                        csv_writer.writerow(line)


class goal_file():
    def __init__(self,archivo_excel,semestre,clave,subject_name):
        self.archivo_excel = archivo_excel
        self.semestre = semestre
        self.clave = clave
        self.subject_name = subject_name
        self.conn = mysql.connector.connect(user="root", password="123456",
                                       host="localhost",
                                       database="fca_materias",
                                       port='3306'
                                       )
        self.cur = self.conn.cursor()

    def progreso(self):
        self.cur.execute("SELECT valor,calificacion,status FROM actividades WHERE usuario = '1' and clave_materia= '" + str(self.clave) +
                         "' AND semestre = '" + self.semestre + "'")

        cal_valor = self.cur.fetchall()
        self.cur.execute("SELECT meta FROM materias_usuario WHERE user_id = '1' and clave_materia= '" + str(self.clave) +
                         "' AND semestre_id = '" + self.semestre + "'")
        meta = self.cur.fetchone()[0]

        resultados = dict()
        acumulado = 0
        max_posible = 0
        for registro in cal_valor:
            valor = registro[0]
            calificacion = registro[1]
            status = registro[2]
            if calificacion != None and status != "Atrasada":
                ponderacion = (calificacion*valor)/10
                acumulado += ponderacion
                max_posible += ponderacion
            elif status == "Atrasada":
                max_posible += 0
            elif status == "Por entregar":
                max_posible += valor
        resultados['acumulado'] = float("{:.2f}".format(acumulado*10))
        resultados['max_posible'] = float("{:.2f}".format(max_posible*10))
        resultados['meta'] = meta
        ## Enviando la suma de calificaciones al libro de excel
        wb = load_workbook(filename=self.archivo_excel,read_only=False,keep_vba=True)
        ws = wb['Hoja1']
        ws['a2'] = self.subject_name[0]
        ws['b2'] = int(self.clave)
        ws['d2'] = acumulado*10
        wb.save(self.archivo_excel)
        app = xw.App(visible=False)
        wb = xw.Book('assets/materia_dashboard_material/meta.xlsm')
        macro1 = wb.macro('modulo.progress')
        macro1()
        wb.save()
        wb.app.quit()

        return resultados


#materias = DB_admin().materias_por_semestre(semestre='23-2',usuario='1')
#print(materias)

#DB_admin(usuario=1).load_data(semestre='23-2',archivo="assets/files/actividades_por_materia/actividades_load.csv")

#DB_admin().materias_por_semestre('23-2',usuario='1')
#archiv = r'C:\Users\ivan_\OneDrive - UNIVERSIDAD NACIONAL AUTÓNOMA DE MÉXICO\Desktop\repositorios\suayedApp\assets\materia_dashboard_material\meta.xlsm'
#goal_file(archiv,archivo,'1343','COMPORTAMIENTO EN LAS ORGANIZACIONES').change_cell()

# obtener_materias(archivo).lista_semester()
# obtener_materias(archivo).define_semester('Semestre 22-2',
# r"C:\Users\ivan_\OneDrive - UNIVERSIDAD NACIONAL AUTÓNOMA DE MÉXICO\Desktop\repositorios\suayedApp\assets\BD/semestres_materias.csv")

# print(obtener_materias(archivo).dias_con_pendientes('05'))
# obtener_materias(archivo).actualizar_estados()
# obtener_materias(r'C:\Python310\PycharmProjects\kivyGUI\virt\KivyMDNavDrawerAndScreenManager\assets\BD\materias.csv').definir_lista('TwoLineRightIconListItem')
# obtener_materias(r'C:\Python310\PycharmProjects\kivyGUI\virt\KivyMDNavDrawerAndScreenManager\assets\BD\materias.csv').total_actividades('PRINCIPIOS Y TECNICAS DE LA INVESTIGACION')

# obtener_materias(r'C:\Python310\PycharmProjects\kivyGUI\virt\KivyMDNavDrawerAndScreenManager\assets\BD\materias.csv').actualizar_DB('COSTOS','Unidad 5 / Actividad complementaria 1 /')

# print(obtener_materias(r'C:\Python310\PycharmProjects\kivyGUI\virt\KivyMDNavDrawerAndScreenManager\assets\BD\materias.csv').subject_actividades('ÉTICA EN LAS ORGANIZACIONES'))

# print(obtener_materias(r'C:\Python310\PycharmProjects\kivyGUI\virt\KivyMDNavDrawerAndScreenManager\assets\BD\materias.csv').total_actividades("COSTOS"))

# obtener_materias(r'C:\Python310\PycharmProjects\kivyGUI\virt\KivyMDNavDrawerAndScreenManager\assets\BD\materias.csv').define_activity('actividad1')

# obtener_materias(r'C:\Python310\PycharmProjects\kivyGUI\virt\KivyMDNavDrawerAndScreenManager\assets\BD\materias.csv').obtener_activity_name()

# obtener_materias(r'C:\Python310\PycharmProjects\kivyGUI\virt\KivyMDNavDrawerAndScreenManager\assets\BD\materias.xlsx').hjh()


# print(estatus_feedback(r'C:\Python310\PycharmProjects\kivyGUI\virt\KivyMDNavDrawerAndScreenManager\assets\BD\materias.csv','PRINCIPIOS Y TECNICAS DE LA INVESTIGACION','U1_Act_aprend_5').resultados())

# obtener_materias(r'C:\Python310\PycharmProjects\kivyGUI\virt\KivyMDNavDrawerAndScreenManager\assets\BD\materias.csv').define_subject('actividad1')

# obtener_materias(r'C:\Python310\PycharmProjects\kivyGUI\virt\KivyMDNavDrawerAndScreenManager\assets\BD\materias.csv').obtener_materia_clave('COSTOS')


# activi = ['Unidad 1 / Actividad complementaria 1 /', 'Unidad 2 / Actividad complementaria 1 /', 'Unidad 2 / Cuestionario de reforzamiento /', 'Unidad 3 / Actividad complementaria 1 /', 'Unidad 4 / Cuestionario de reforzamiento /', 'Unidad 4 / Actividad complementaria 1/', 'Unidad 5 / Actividad complementaria 1 /', 'Unidad 6 / Cuestionario de reforzamiento /', 'Unidad 6 / Actividad complementaria 1 /', 'Unidad 7 / Cuestionario de reforzamiento /', 'Unidad 8 / Actividad complementaria 1 /', 'Unidad 1 / Cuestionario de reforzamiento /', 'Unidad 1 / Actividad complementaria 1 /']

#
# vaciar_feedback(r'C:\Python310\PycharmProjects\kivyGUI\virt\KivyMDNavDrawerAndScreenManager\assets\BD\materias-copia.csv','COMPORTAMIENTO EN LAS ORGANIZACIONES',retro).vaciar_resultados()

# btener_materias(r'C:\Python310\PycharmProjects\kivyGUI\virt\KivyMDNavDrawerAndScreenManager\assets\BD\materias.csv').define_activity('fefef')


# obtener_materias(r'C:\Python310\PycharmProjects\kivyGUI\virt\KivyMDNavDrawerAndScreenManager\assets\BD\materias.csv').total_actividades('DESARROLLO SUSTENTABLE Y LAS ORGANIZACIONES')
