import csv
import datetime
from datetime import date
import chardet
import pandas as pd
import xlwings as xw
from openpyxl import load_workbook
import calendar
import mysql.connector


class obtener_materias():

    def __init__(self, archivo_materias):
        self.archivo_materias = archivo_materias
        self.conn = mysql.connector.connect(user="root", password="123456",
                                       host="localhost",
                                       database="fca_materias",
                                       port='3306'
                                       )
        self.cur = self.conn.cursor()
    def lista_semester(self):  # Devuelve una lista con los semestres disponibles

        self.cur.execute(
            '''
                SELECT DISTINCT name FROM semestres
            '''
        )
        semestres = self.cur.fetchall()
        semestres_ = list()
        for semestre in semestres:
            semestres_.append(semestre[0])
        return semestres_


    def define_semester(self, semestre, archivo_aux):
        '''SET THE SEMESTER WE'RE GONNA WORK WITH IN USER_SETTINGS TABLE'''
        self.cur.execute("SELECT id FROM semestres WHERE name = '" + semestre + "'")  ## busca el id del semestre que se seleccionó
        semestre = self.cur.fetchone()[0]

        self.cur.execute('''
            UPDATE user_settings SET tipo_lista = 'TwoLineRightIconListItem' WHERE user_id = 1 
        ''')
        self.cur.execute('''
            UPDATE user_settings SET status = 'Por entregar' WHERE user_id = 1
        ''')
        self.cur.execute("UPDATE user_settings SET semestre_sel = '" + str(semestre) + "' WHERE user_id = 1") ## establece el semestre seleccionado en la tabla user_settings
        subject_name = self.obtener_materia_name_(modo=2)
        self.cur.execute("UPDATE user_settings SET materia_sel = '" + subject_name[0] + "' WHERE user_id = 1") ## establece la primera materia econtrada del semestre establecido
        self.conn.commit()

    def dias_con_pendientes(self, modo, month=None, año=None, fecha=None):
        if modo == 1:  ## Regresa una lista con los días que tienen pendientes
            primer_dia_mes = año + '-' + str(month).zfill(2) +'-' + '01'
            ultimo_dia_mes = año + '-' + str(month).zfill(2) + '-' + str(calendar.monthrange(int(año), month)[1])

            fechas_mes = "SELECT fecha_entrega FROM actividades " \
                                  "WHERE fecha_entrega BETWEEN '" + primer_dia_mes + "' AND '" + ultimo_dia_mes + "'"

            self.cur.execute(fechas_mes)
            fechas_ = self.cur.fetchall()
            lista_dias = list()
            if len(fechas_)!=0:
                fechas,= list(zip(*fechas_))
                for fecha in fechas:
                    fecha = fecha.strftime("%Y-%m-%d").split("-")[2]
                    lista_dias.append(fecha)

            return lista_dias

        elif modo ==2:  ## Regresa una lista con el nomnbre de las materias, act y status
            dia = "SELECT name,clave_materia FROM actividades " \
                                  "WHERE fecha_entrega = '" + fecha + "'"

            self.cur.execute(dia)
            registros = self.cur.fetchall()
            materias = list()
            actividades = list()
            status = list()
            materia_actividad = list()
            for registro in registros:
                clave = registro[1]
                acti_name = registro[0]
                acti_name = acti_name.replace(" / ", "_").replace("Unidad ", "U").replace("/", "").replace(
                    "Actividad complementaria", "Act_compl"). \
                    replace("Cuestionario de reforzamiento", "Cuest_refor").replace("Actividad", "Act")
                actividades.append(acti_name)
                materia_nombre = '''SELECT materia FROM materias_fca  
                                    JOIN actividades 
                                    ON materias_fca.clave = actividades.clave_materia
                                    where materias_fca.clave =''' + str(clave) + ''' LIMIT 1'''
                self.cur.execute(materia_nombre)
                status_ = self.cur.fetchone()
                status.append(status_[0])


                self.cur.execute("SELECT status FROM actividades where clave_materia ='" + str(clave) + "' AND name= '" + registro[0] + "'")
                materia_nombre = self.cur.fetchone()
                materias.append(materia_nombre[0])

            materia_actividad.append(materias)
            materia_actividad.append(actividades)
            materia_actividad.append(status)
            return materia_actividad

    def remember_status(self,modo):  ##Esta función juega con el remember_status
        remember_status = ('''
            SELECT recordar_inicio_sesion FROM user_settings WHERE user_id = 1
        '''
        )
        self.cur.execute(remember_status)
        remember_status = self.cur.fetchone()[0]

        if modo ==1: ## Se presionó el butón


            if remember_status == 1:
                self.cur.execute('''
                    UPDATE user_settings SET recordar_inicio_sesion=0 WHERE user_id = 1
                ''')
            else:
                self.cur.execute('''
                    UPDATE user_settings SET recordar_inicio_sesion=1 WHERE user_id = 1
                ''')

            self.conn.commit()
        else: #pura consulta
            return remember_status


    def extracción_materias(self):
        '''Devuelve las materias conforme al semestre seleccionado'''
        semestre = "SELECT semestre_sel FROM user_settings WHERE user_id = 1"
        self.cur.execute(semestre)
        semestre = self.cur.fetchone()[0]


        materias = "SELECT DISTINCT materia FROM materias_fca JOIN actividades ON materias_fca.clave = actividades.clave_materia WHERE actividades.semestre ='" + semestre + "'"
        self.cur.execute(materias)
        materias = self.cur.fetchall()
        materias_ = list()
        for materia in materias:
            materias_.append(materia[0])
        return materias_

    def status(self,modo,materia=None,status_=None):
        if modo == 1:  ## Actualiza solo el estado
            act_estatus = self.cur.execute("SELECT name,date_format(fecha_entrega,'%d-%m-%Y'),date_format(entregada_el,'%d-%m-%Y'),status FROM actividades")
            act_status = [list(i) for i in self.cur.fetchall()]

            i = 0
            for status in act_status:
                fecha_entrega = status[1]
                entregada_el = status[2]
                fecha_entrega = datetime.datetime.strptime(fecha_entrega, '%d-%m-%Y')
                if entregada_el == None:
                    status_ = 'Por entregar' if datetime.datetime.today() < fecha_entrega else 'Atrasada'
                else:
                    entregada_el = datetime.datetime.strptime(entregada_el, '%d-%m-%Y')
                    status_ = 'Entregada a tiempo' if entregada_el<=fecha_entrega else 'Entregada con atraso'
                i+=1

                self.cur.execute("UPDATE actividades SET status = '" +status_ + "' WHERE id ='" + str(i) +"'")
                self.conn.commit()
        elif modo ==2: ## Trae las actividaddes de las materias con el estado dado
            act_estatus = self.cur.execute(
                "SELECT name,date_format(fecha_entrega,'%d-%m-%Y'),date_format(entregada_el,'%d-%m-%Y'),status FROM actividades WHERE clave_materia= '" + str(materia) + "' AND status = '" + status_ +"'")
            act_status = [list(i) for i in self.cur.fetchall()]
            activity_date = dict()
            for act in act_status:
                act[0] = act[0].replace(" / ", "_").replace("Unidad ", "U").replace("/", "").replace("Actividad complementaria","Act_compl"). \
                replace("Cuestionario de reforzamiento","Cuest_refor").replace("Actividad","Act")
                actividad = act[0]
                fecha_entrega = act[1]
                entregada_el = act[2]
                fecha_entrega = datetime.datetime.strptime(fecha_entrega, '%d-%m-%Y')
                if status_ == 'Por entregar':
                    activity_date[actividad] = f'Due to {fecha_entrega}'
                    self.cur.execute('''
                        UPDATE user_settings SET tipo_lista = 'TwoLineRightIconListItem' WHERE user_id = 1
                    ''')
                elif status_ == 'Entregada a tiempo':
                    self.cur.execute('''
                        UPDATE user_settings SET tipo_lista = 'TwoLineListItem' WHERE user_id = 1
                    ''')
                    entregada_el = datetime.datetime.strptime(entregada_el, '%d-%m-%Y')
                    activity_date[actividad] = f'Delivered {(datetime.datetime.today()-entregada_el).days} days ago'
                elif status_ == 'Entregada con atraso':
                    self.cur.execute('''
                        UPDATE user_settings SET tipo_lista = 'TwoLineListItem' WHERE user_id = 1
                    ''')
                    entregada_el = datetime.datetime.strptime(entregada_el, '%d-%m-%Y')
                    activity_date[actividad] = f'Delivered {(datetime.datetime.today()-entregada_el).days} days ago'
                elif status_ == 'Atrasada':
                    self.cur.execute('''
                        UPDATE user_settings SET tipo_lista = 'TwoLineRightIconListItem' WHERE user_id = 1
                    ''')
                    activity_date[actividad] = f'This activity was due for {(fecha_entrega)} days delayed'
            self.conn.commit()
            return activity_date
        elif modo ==3: ## Busca cual es el estado en user_settings
            self.cur.execute("SELECT status FROM user_settings WHERE user_id = 1")
            status = self.cur.fetchone()[0]
            return status

        elif modo == 4: ## Cambia el estado en user_setting según lo seleccionado
            self.cur.execute('''
                UPDATE user_settings SET status = "'''     + status_  + '''" WHERE user_id = 1
            ''')
            self.conn.commit()

        elif modo ==5:
            actividades = self.cur.execute(
                "SELECT name FROM actividades WHERE clave_materia= '" + str(materia) + "'")

            actividades = [i[0] for i in self.cur.fetchall()]
            return actividades


    def define_activity(self, actividad):
        '''Estblece en la tabla user_settings el valor de la act seleccionada en la pantalla 2'''
        if "Act_compl" in actividad:
            actividad = actividad.replace("U", "Unidad ").replace(
                "Act_compl", "Actividad complementaria").replace("_", " / ") + "/"
        else:
            actividad = actividad.replace("U", "Unidad ").replace(
                "Act_compl", "Actividad complementaria"). \
                replace("Cuest_refor", "Cuestionario de reforzamiento").replace("Act", "Actividad").replace("_", " / ") + "/"
        self.cur.execute('''
            UPDATE user_settings SET act_sel = "''' + actividad + '''" WHERE user_id = 1
        ''')
        self.conn.commit()

    def estado_actividad(self, clave, actividad):
        if "Act_compl" in actividad:
            actividad = actividad.replace("U", "Unidad ").replace(
                "Act_compl", "Actividad complementaria").replace("_", " / ") + "/"
        else:
            actividad = actividad.replace("U", "Unidad ").replace(
                "Act_compl", "Actividad complementaria"). \
                replace("Cuest_refor", "Cuestionario de reforzamiento").replace("Act", "Actividad").replace("_", " / ") + "/"
        '''Busca los valores del feedback de la actividad seleccionada en la p2 y los muestra en la pantalla de feedback'''
        self.cur.execute(
            "SELECT date_format(entregada_el,'%d-%m-%Y'),valor,status,calificacion,calificada_en,comentarios FROM actividades WHERE clave_materia= '" + str(clave) + "' AND name = '" + actividad + "'")
        feedback = self.cur.fetchone()
        feedback_ = list()
        feedback = [elem if elem != None else "" for elem in feedback]
        entregada_el = feedback[0]
        valor = feedback[1]
        status = feedback[2]
        cal = feedback[3]
        calificada_en = feedback[4]
        comentarios = feedback[5]
        feedback_.append(entregada_el)
        feedback_.append(valor)
        feedback_.append(status)
        feedback_.append(cal)
        feedback_.append(calificada_en)
        feedback_.append(comentarios)
        return feedback_

    def obtener_materia_name_(self,modo,subject_name_=None):
        if modo ==1: ## Busca el nombre de la materia
            self.cur.execute("SELECT materia_sel FROM user_settings WHERE user_id = 1")
            subject_name = self.cur.fetchone()
            return subject_name
        elif modo ==2: ## Seleccionad el primer nombre de la materia que encuentre de acuerdo al semestre
            self.cur.execute( '''SELECT DISTINCT materia FROM materias_fca
                                JOIN actividades  
                                ON materias_fca.clave = actividades.clave_materia   LIMIT 1''')
            subject_name = self.cur.fetchone()
            return subject_name
        elif modo ==3:   ## Busca la clave de la materia
            self.cur.execute( '''
            SELECT DISTINCT clave FROM materias_fca
            JOIN actividades
            ON materias_fca.clave = actividades.clave_materia
            WHERE materia = "''' + subject_name_ + '''"''')
            clave = self.cur.fetchone()
            return clave

        elif modo == 4: ## Busca el nombre de la actividdad seleccionada en la pantalla 2
            self.cur.execute("SELECT act_sel FROM user_settings WHERE user_id = 1")
            acti_name = self.cur.fetchone()
            acti_name = acti_name[0]
            acti_name = acti_name.replace(" / ", "_").replace("Unidad ", "U").replace("/", "").replace(
                "Actividad complementaria", "Act_compl"). \
                replace("Cuestionario de reforzamiento", "Cuest_refor").replace("Actividad", "Act")
            return acti_name

    def list_type(self):
        self.cur.execute("SELECT tipo_lista FROM user_settings WHERE user_id = 1")
        tipo_lista = self.cur.fetchone()
        return tipo_lista[0]

    def obtener_materia_grupo(self, clave):
        self.cur.execute("SELECT DISTINCT grupo FROM actividades WHERE clave_materia= '" + str(
            clave)  + "'")
        grupo = self.cur.fetchone()[0]
        return grupo

    def actualizar_DB(self, materia, actividad):
        self.cur.execute("SELECT date_format(fecha_entrega,'%d-%m-%Y') FROM actividades WHERE clave_materia= '" + str(
            materia) + "' AND name = '" + actividad + "'")
        datos_act = self.cur.fetchone()

        hoy = datetime.datetime.today()
        self.cur.execute("UPDATE actividades SET entregada_el = '" + datetime.datetime.strftime(hoy,"%Y-%m-%d") + "' WHERE clave_materia= '" + str(
            materia) + "' AND name = '" + actividad + "'")

        fecha_entrega = datos_act[0]
        fecha_entrega = datetime.datetime.strptime(fecha_entrega,"%d-%m-%Y")
        if hoy > fecha_entrega:
            estado = "Entregada con atraso"
        else:
            estado = "Entregada a tiempo"
        self.cur.execute("UPDATE actividades SET status = '" + estado + "' WHERE clave_materia= '" + str(
            materia) + "' AND name = '" + actividad + "'")
        self.conn.commit()
    def condensado_tareas(self,clave):
        semestre =self.semestre_seleccionado()
        act_valor_cali = list()
        d = ("SELECT name,valor,calificacion FROM actividades WHERE usuario = 1 AND semestre = '{}' AND clave_materia = '{}'").format(str(semestre),str(clave))
        self.cur.execute(d)
        table_rows = self.cur.fetchall()
        df = pd.read_sql(d,con = self.conn)

        char_to_replace = {" / ":"_",
                           "Unidad ": "U",
                           "Actividad complementaria": "Act_compl",
                           "Cuestionario de reforzamiento": "Cuest_refor",
                            "Actividad": "Act"}
        for key,value in char_to_replace.items():
            df['name'] = df['name'].str.replace(key,value)
        df.rename(columns={'name':'Act','valor':'Valor','calificacion':'Calificación'},inplace=True)
        return df

    def semestre_seleccionado(self):
        semestre = "SELECT semestre_sel FROM user_settings WHERE user_id = 1"
        self.cur.execute(semestre)
        semestre = self.cur.fetchone()[0]
        return semestre


class estatus_feedback():
    def __init__(self, archivo_materias, materia, actividad):
        self.archivo_materias = archivo_materias
        self.actividad = actividad
        self.materia = materia

    def resultados(self):
        with open(self.archivo_materias, 'r') as file:
            reader = csv.reader(file)
            activity_status = list()
            for line in reader:

                if line[1] == self.materia and line[2] == self.actividad:
                    fecha_entregada = line[4]
                    status = line[5]
                    calificacion = line[6]
                    calificada_el = line[7]
                    comentarios = line[8]
                    valor = line[13]

                    activity_status.append(fecha_entregada)
                    activity_status.append(status)
                    activity_status.append(calificacion)
                    activity_status.append(calificada_el)
                    activity_status.append(comentarios)
                    activity_status.append(valor)
                    break

            return activity_status


class vaciar_feedback():
    def __init__(self, archivo_materias, materia, actividades_feedback):
        self.archivo_materias = archivo_materias
        self.materia = materia
        self.actividades_feedback = actividades_feedback

    def vaciar_resultados(self):
        with open(self.archivo_materias, 'r') as file:

            reader = csv.reader(file)
            myList = list(reader)
            for line in reader:
                myList.append(line)
            for k, v in self.actividades_feedback.items():
                c = 0
                for row in myList:
                    if row[3] == self.materia and row[4] == k:
                        if v[0]!="" and v[0]!="Sin envío":
                            v[0] = float(v[0].split("/")[0])
                        else:pass
                        myList[c][8] = v[0]  # calificacion
                        myList[c][9] = v[1]  # calificada el
                        myList[c][10] = v[2]  # comentarios
                    c += 1
                    continue

                with open(self.archivo_materias, 'rb') as rawdata:
                    result = chardet.detect(rawdata.read(100000))
                my_new_list = open(self.archivo_materias, 'w', newline='', encoding=result['encoding'])
                csv_writer = csv.writer(my_new_list)
                for line in myList:
                    try:
                        csv_writer.writerow(line)
                    except:
                        line[10] = line[10].encode('utf8').decode('ascii', 'ignore')
                        csv_writer.writerow(line)


class goal_file():
    def __init__(self,archivo_excel,semestre,clave,subject_name):
        self.archivo_excel = archivo_excel
        self.semestre = semestre
        self.clave = clave
        self.subject_name = subject_name
        self.conn = mysql.connector.connect(user="root", password="123456",
                                       host="localhost",
                                       database="fca_materias",
                                       port='3306'
                                       )
        self.cur = self.conn.cursor()

    def progreso(self):
        self.cur.execute("SELECT valor,calificacion FROM actividades WHERE usuario = '1' and clave_materia= '" + str(self.clave) +
                         "' AND semestre = '" + self.semestre + "'")

        cal_valor = self.cur.fetchall()
        self.cur.execute("SELECT meta FROM materias_usuario WHERE user_id = '1' and clave_materia= '" + str(self.clave) +
                         "' AND semestre_id = '" + self.semestre + "'")
        print("SELECT meta FROM materias_usuario WHERE user_id = '1' and clave_materia= '" + str(self.clave) +
                         "' AND semestre_id = '" + self.semestre + "'")
        meta = self.cur.fetchone()[0]

        resultados = dict()
        acumulado = 0
        max_posible = 0
        for registro in cal_valor:
            valor = registro[0]
            calificacion = registro[1]
            if calificacion != None:
                ponderacion = (calificacion*valor)/10
                acumulado += ponderacion
                max_posible += ponderacion
            else:
                max_posible += valor
        resultados['acumulado'] = float("{:.2f}".format(acumulado*10))
        resultados['max_posible'] = float("{:.2f}".format(max_posible*10))
        resultados['meta'] = meta
        ## Enviando la suma de calificaciones al libro de excel
        wb = load_workbook(filename=self.archivo_excel,read_only=False,keep_vba=True)
        ws = wb['Hoja1']
        ws['a2'] = self.subject_name[0]
        ws['b2'] = int(self.clave)
        ws['d2'] = acumulado*10
        wb.save(self.archivo_excel)
        app = xw.App(visible=False)
        wb = xw.Book('assests/materia_dashboard_material/meta.xlsm')
        macro1 = wb.macro('modulo.progress')
        macro1()
        wb.save()
        wb.app.quit()

        return resultados


archivo = r'C:\Users\ivan_\OneDrive - UNIVERSIDAD NACIONAL AUTÓNOMA DE MÉXICO\Desktop\repositorios\suayedApp\assests\BD\materias.csv'
obtener_materias(archivo).status(modo=2,materia='1255',status_='Entregada con atraso')

#archiv = r'C:\Users\ivan_\OneDrive - UNIVERSIDAD NACIONAL AUTÓNOMA DE MÉXICO\Desktop\repositorios\suayedApp\assests\materia_dashboard_material\meta.xlsm'
#goal_file(archiv,archivo,'1343','COMPORTAMIENTO EN LAS ORGANIZACIONES').change_cell()

# obtener_materias(archivo).lista_semester()
# obtener_materias(archivo).define_semester('Semestre 22-2',
# r"C:\Users\ivan_\OneDrive - UNIVERSIDAD NACIONAL AUTÓNOMA DE MÉXICO\Desktop\repositorios\suayedApp\assests\BD/semestres_materias.csv")

# print(obtener_materias(archivo).dias_con_pendientes('05'))
# obtener_materias(archivo).actualizar_estados()
# obtener_materias(r'C:\Python310\PycharmProjects\kivyGUI\virt\KivyMDNavDrawerAndScreenManager\assests\BD\materias.csv').definir_lista('TwoLineRightIconListItem')
# obtener_materias(r'C:\Python310\PycharmProjects\kivyGUI\virt\KivyMDNavDrawerAndScreenManager\assests\BD\materias.csv').total_actividades('PRINCIPIOS Y TECNICAS DE LA INVESTIGACION')

# obtener_materias(r'C:\Python310\PycharmProjects\kivyGUI\virt\KivyMDNavDrawerAndScreenManager\assests\BD\materias.csv').actualizar_DB('COSTOS','Unidad 5 / Actividad complementaria 1 /')

# print(obtener_materias(r'C:\Python310\PycharmProjects\kivyGUI\virt\KivyMDNavDrawerAndScreenManager\assests\BD\materias.csv').subject_actividades('ÉTICA EN LAS ORGANIZACIONES'))

# print(obtener_materias(r'C:\Python310\PycharmProjects\kivyGUI\virt\KivyMDNavDrawerAndScreenManager\assests\BD\materias.csv').total_actividades("COSTOS"))

# obtener_materias(r'C:\Python310\PycharmProjects\kivyGUI\virt\KivyMDNavDrawerAndScreenManager\assests\BD\materias.csv').define_activity('actividad1')

# obtener_materias(r'C:\Python310\PycharmProjects\kivyGUI\virt\KivyMDNavDrawerAndScreenManager\assests\BD\materias.csv').obtener_activity_name()

# obtener_materias(r'C:\Python310\PycharmProjects\kivyGUI\virt\KivyMDNavDrawerAndScreenManager\assests\BD\materias.xlsx').hjh()


# print(estatus_feedback(r'C:\Python310\PycharmProjects\kivyGUI\virt\KivyMDNavDrawerAndScreenManager\assests\BD\materias.csv','PRINCIPIOS Y TECNICAS DE LA INVESTIGACION','U1_Act_aprend_5').resultados())

# obtener_materias(r'C:\Python310\PycharmProjects\kivyGUI\virt\KivyMDNavDrawerAndScreenManager\assests\BD\materias.csv').define_subject('actividad1')

# obtener_materias(r'C:\Python310\PycharmProjects\kivyGUI\virt\KivyMDNavDrawerAndScreenManager\assests\BD\materias.csv').obtener_materia_clave('COSTOS')


# activi = ['Unidad 1 / Actividad complementaria 1 /', 'Unidad 2 / Actividad complementaria 1 /', 'Unidad 2 / Cuestionario de reforzamiento /', 'Unidad 3 / Actividad complementaria 1 /', 'Unidad 4 / Cuestionario de reforzamiento /', 'Unidad 4 / Actividad complementaria 1/', 'Unidad 5 / Actividad complementaria 1 /', 'Unidad 6 / Cuestionario de reforzamiento /', 'Unidad 6 / Actividad complementaria 1 /', 'Unidad 7 / Cuestionario de reforzamiento /', 'Unidad 8 / Actividad complementaria 1 /', 'Unidad 1 / Cuestionario de reforzamiento /', 'Unidad 1 / Actividad complementaria 1 /']

#
# vaciar_feedback(r'C:\Python310\PycharmProjects\kivyGUI\virt\KivyMDNavDrawerAndScreenManager\assests\BD\materias-copia.csv','COMPORTAMIENTO EN LAS ORGANIZACIONES',retro).vaciar_resultados()

# btener_materias(r'C:\Python310\PycharmProjects\kivyGUI\virt\KivyMDNavDrawerAndScreenManager\assests\BD\materias.csv').define_activity('fefef')


# obtener_materias(r'C:\Python310\PycharmProjects\kivyGUI\virt\KivyMDNavDrawerAndScreenManager\assests\BD\materias.csv').total_actividades('DESARROLLO SUSTENTABLE Y LAS ORGANIZACIONES')
